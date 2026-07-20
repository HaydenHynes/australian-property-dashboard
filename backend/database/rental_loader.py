"""Load and aggregate NSW Fair Trading rental-bond workbooks."""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from backend.database.connection import Database
from scripts.download.download_rental_bonds import SOURCE_PAGE, WORKBOOKS

REFRESH_SUBURB_POSTCODES_SQL = """
TRUNCATE suburb_postcodes;
INSERT INTO suburb_postcodes (locality, postcode, supporting_sales)
SELECT
    UPPER(TRIM(property_locality)),
    LPAD(TRIM(property_postcode), 4, '0'),
    COUNT(*)
FROM property_sales
WHERE property_locality IS NOT NULL
  AND property_postcode ~ '^[0-9]{4}$'
GROUP BY UPPER(TRIM(property_locality)), LPAD(TRIM(property_postcode), 4, '0');
"""

REFRESH_MONTHLY_MARKET_SQL = """
TRUNCATE rental_market_monthly;
INSERT INTO rental_market_monthly (
    month,
    postcode,
    dwelling_type,
    bedrooms,
    median_weekly_rent,
    lower_quartile_rent,
    upper_quartile_rent,
    lodgement_count
)
SELECT
    DATE_TRUNC('month', lodgement_date)::date,
    postcode,
    COALESCE(dwelling_type, 'U'),
    COALESCE(bedrooms, -1),
    PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY weekly_rent),
    PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY weekly_rent),
    PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY weekly_rent),
    COUNT(*)
FROM rental_bond_lodgements
WHERE weekly_rent BETWEEN 50 AND 5000
  AND dwelling_type IN ('F', 'H', 'T', 'O', 'U')
GROUP BY
    DATE_TRUNC('month', lodgement_date),
    postcode,
    COALESCE(dwelling_type, 'U'),
    COALESCE(bedrooms, -1);
"""


def _normalise_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _normalise_postcode(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip().split(".")[0].zfill(4)
    return text if len(text) == 4 and text.isdigit() else None


def _normalise_rent(value) -> Decimal | None:
    try:
        rent = Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return None
    return rent if Decimal("1") <= rent <= Decimal("50000") else None


def _normalise_bedrooms(value) -> int | None:
    try:
        bedrooms = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return bedrooms if 0 <= bedrooms <= 20 else None


def _find_data_sheet(workbook):
    for sheet in workbook.worksheets:
        headers = [sheet.cell(3, column).value for column in range(1, 6)]
        if headers == [
            "Lodgement Date",
            "Postcode",
            "Dwelling Type",
            "Bedrooms",
            "Weekly Rent",
        ]:
            return sheet
    raise ValueError("Rental-bond data sheet with expected headers was not found")


def load_rental_workbook(db: Database, path: Path, source_url: str) -> int:
    """Load one workbook idempotently and return inserted row count."""
    checksum = hashlib.sha256(path.read_bytes()).hexdigest()
    with db.connection() as conn:
        existing = conn.execute(
            "SELECT sha256, row_count FROM rental_data_imports WHERE source_file = %s",
            (path.name,),
        ).fetchone()
        if existing:
            if existing[0] != checksum:
                raise ValueError(f"Source file changed since import: {path.name}")
            return existing[1]

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = _find_data_sheet(workbook)
    valid_rows: list[tuple] = []
    for source_row, row in enumerate(
        sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        if len(row) < 5:
            continue
        lodgement_date = _normalise_date(row[0])
        postcode = _normalise_postcode(row[1])
        weekly_rent = _normalise_rent(row[4])
        if not lodgement_date or not postcode or weekly_rent is None:
            continue
        dwelling_type = str(row[2]).strip().upper()[:1] if row[2] else "U"
        if dwelling_type not in {"F", "H", "T", "O", "U"}:
            dwelling_type = "U"
        valid_rows.append(
            (
                path.name,
                source_row,
                lodgement_date,
                postcode,
                dwelling_type,
                _normalise_bedrooms(row[3]),
                weekly_rent,
            )
        )
    workbook.close()

    with db.connection() as conn:
        conn.execute(
            """
            INSERT INTO rental_data_imports (source_file, source_url, sha256, row_count)
            VALUES (%s, %s, %s, %s)
            """,
            (path.name, source_url, checksum, len(valid_rows)),
        )
        with conn.cursor() as cursor:
            with cursor.copy(
                """
                COPY rental_bond_lodgements (
                    source_file, source_row, lodgement_date, postcode,
                    dwelling_type, bedrooms, weekly_rent
                ) FROM STDIN
                """,
            ) as copy:
                for row in valid_rows:
                    copy.write_row(row)
        conn.commit()
    return len(valid_rows)


def refresh_rental_aggregates(db: Database) -> None:
    with db.connection() as conn:
        conn.execute(REFRESH_SUBURB_POSTCODES_SQL)
        conn.execute(REFRESH_MONTHLY_MARKET_SQL)
        conn.commit()


def load_all_rental_workbooks(db: Database, directory: Path) -> int:
    total = 0
    for filename, url in WORKBOOKS.items():
        path = directory / filename
        if not path.exists():
            raise FileNotFoundError(f"Missing rental workbook: {path}")
        rows = load_rental_workbook(db, path, url)
        total += rows
        print(f"Loaded {rows:,} rows from {filename}")
    refresh_rental_aggregates(db)
    print(f"Source: {SOURCE_PAGE}")
    return total
